import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('table1.xlsx') # this path searches from where the terminal is
sheet1 = wb['Sheet1']

cell = sheet1['a1']
cell = sheet1.cell(1, 1) # same as above

# print(cell.value)

for row in range(2, sheet1.max_row + 1):
    cell = sheet1.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet1.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet1, # all the values in the D column
          min_row=2,
          max_row=sheet1.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet1.add_chart(chart, 'e2')

wb.save('table2.xlsx')